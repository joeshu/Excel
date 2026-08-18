from __future__ import annotations

import csv
import json
from datetime import date, timedelta
from pathlib import Path
from random import Random

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "sample_data" / "scenarios"
SEED = 20260818


def style_sheet(sheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor="17324D")
        cell.alignment = Alignment(horizontal="center")
    for column in sheet.columns:
        letter = column[0].column_letter
        sheet.column_dimensions[letter].width = max(12, min(24, max(len(str(cell.value or "")) for cell in column) + 2))


def save_source(name: str, headers: list[str], rows: list[list[object]], csv_copy: bool = False) -> dict:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据源"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    style_sheet(sheet)
    source_path = OUTPUT / f"{name}_source.xlsx"
    workbook.save(source_path)
    result = {"path": source_path.name, "rows": len(rows), "fields": headers}
    if csv_copy:
        csv_path = OUTPUT / f"{name}_source.csv"
        with csv_path.open("w", newline="", encoding="utf-8-sig") as target:
            writer = csv.writer(target)
            writer.writerow(headers)
            writer.writerows(rows)
        result["csv_path"] = csv_path.name
    return result


def save_template(name: str, headers: list[str], rows: list[list[object]], formulas: dict[str, str] | None = None, extra_sheets: dict[str, list[list[object]]] | None = None) -> str:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据模板"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    for column, formula in (formulas or {}).items():
        for row_number in range(2, len(rows) + 2):
            sheet[f"{column}{row_number}"] = formula.format(row=row_number)
    style_sheet(sheet)
    for extra_name, extra_rows in (extra_sheets or {}).items():
        extra = workbook.create_sheet(extra_name)
        for row in extra_rows:
            extra.append(row)
        style_sheet(extra)
    path = OUTPUT / f"{name}_template.xlsx"
    workbook.save(path)
    return path.name


def build_scenarios() -> list[dict]:
    random = Random(SEED)
    regions = ["华东", "华南", "华北", "西南"]
    categories = ["硬件", "软件", "服务"]
    scenarios = []

    basic_headers = ["customer_id", "customer_name", "region", "level", "signup_date", "active"]
    basic_rows = [[index, f"客户-{index:03d}", regions[index % len(regions)], "重点" if index % 3 == 0 else "普通", date(2024, 1, 1) + timedelta(days=index), index % 4 != 0] for index in range(1, 13)]
    basic_mapping = {f"数据模板!{column}": field for column, field in zip("ABCDEF", basic_headers)}
    scenarios.append({"name": "basic_no_formula", "mode": "formula", "complexity": "basic", "source": save_source("basic_no_formula", basic_headers, basic_rows), "template": save_template("basic_no_formula", basic_headers, [[None] * len(basic_headers) for _ in basic_rows]), "mapping": basic_mapping, "formula_count": 0})

    formula_headers = ["order_id", "region", "category", "quantity", "unit_price", "discount_rate", "gross_amount", "net_amount", "risk"]
    formula_rows = [[f"SO-{index:04d}", regions[index % len(regions)], categories[index % len(categories)], index % 9 + 1, round(50 + random.random() * 800, 2), round((index % 5) * 0.05, 2), None, None, None] for index in range(1, 31)]
    formula_mapping = {f"数据模板!{column}": field for column, field in zip("ABCDEF", formula_headers[:6])}
    formula_template = save_template("standard_formula", formula_headers, [[None] * len(formula_headers) for _ in formula_rows], {"G": "=D{row}*E{row}", "H": "=G{row}*(1-F{row})", "I": '=IF(H{row}>2000,"复核","正常")'})
    scenarios.append({"name": "standard_formula", "mode": "formula", "complexity": "medium", "source": save_source("standard_formula", formula_headers, formula_rows), "template": formula_template, "mapping": formula_mapping, "formula_count": 90})

    complex_headers = ["record_id", "order_no", "region", "category", "quantity", "unit_price", "discount_rate", "owner", "amount", "status"]
    complex_rows = [[index, f"CX-{index:05d}", regions[index % len(regions)], categories[index % len(categories)], 0 if index % 17 == 0 else index % 120 + 1, round(20 + random.random() * 1200, 2), None if index % 19 == 0 else round((index % 6) * 0.03, 2), f"负责人-{index % 8 + 1}", None, "待复核" if index % 17 == 0 else "正常"] for index in range(1, 121)]
    complex_mapping = {f"数据模板!{column}": field for column, field in zip("ABCDEFGHJ", [*complex_headers[:8], "status"])}
    lookup = [["region", "region_level"], *[[region, "核心区域" if index % 2 else "成长区域"] for index, region in enumerate(regions, start=1)]]
    complex_template = save_template("multi_sheet_complex", complex_headers, [[None] * len(complex_headers) for _ in complex_rows], {"I": "=E{row}*F{row}*(1-IFERROR(G{row},0))"}, {"区域字典": lookup, "汇总": [["region", "net_total", "level"], *[[region, f'=SUMIF(数据模板!C$2:C$121,A{index + 1},数据模板!I$2:I$121)', f'=IFERROR(VLOOKUP(A{index + 1},区域字典!$A:$B,2,FALSE),"未知")'] for index, region in enumerate(regions)]]})
    scenarios.append({"name": "multi_sheet_complex", "mode": "formula", "complexity": "complex", "source": save_source("multi_sheet_complex", complex_headers, complex_rows), "template": complex_template, "mapping": complex_mapping, "formula_count": 128})

    quality_headers = ["id", "name", "amount", "region", "event_date", "remark"]
    quality_rows = [[index, f"异常-{index:03d}", "bad" if index % 7 == 0 else (None if index % 5 == 0 else index * 10), None if index % 4 == 0 else regions[index % len(regions)], None if index % 6 == 0 else date(2025, 3, 1) + timedelta(days=index), "类型混合" if index % 7 == 0 else ""] for index in range(1, 41)]
    quality_mapping = {f"数据模板!{column}": field for column, field in zip("ABCDEF", quality_headers)}
    scenarios.append({"name": "quality_edge_cases", "mode": "formula", "complexity": "edge", "source": save_source("quality_edge_cases", quality_headers, quality_rows, csv_copy=True), "template": save_template("quality_edge_cases", quality_headers, [[None] * len(quality_headers) for _ in quality_rows]), "mapping": quality_mapping, "formula_count": 0})
    return scenarios


def main() -> None:
    OUTPUT.mkdir(parents=True, exist_ok=True)
    manifest = {"seed": SEED, "scenarios": build_scenarios()}
    (OUTPUT / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(f"generated {len(manifest['scenarios'])} scenarios in {OUTPUT}")


if __name__ == "__main__":
    main()
