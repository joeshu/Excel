from datetime import date, timedelta
from pathlib import Path
from random import Random

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "sample_data"
ROWS = 200
SEED = 20260818


def build_workbook() -> tuple[Workbook, Workbook]:
    random = Random(SEED)
    regions = ["华东", "华南", "华北", "西南", "西北"]
    categories = ["硬件", "软件", "服务", "耗材"]
    customer_levels = ["战略客户", "重点客户", "普通客户"]
    headers = [
        "record_id", "order_no", "customer_name", "region", "category",
        "order_date", "quantity", "unit_price", "discount_rate", "sales_owner",
        "gross_amount", "discount_amount", "net_amount", "risk_level", "remark",
    ]
    source = Workbook()
    source_sheet = source.active
    source_sheet.title = "基础数据"
    source_sheet.append(headers)
    for index in range(1, ROWS + 1):
        region = regions[index % len(regions)]
        category = categories[index % len(categories)]
        quantity = (index * 7) % 120 + 1
        if index % 17 == 0:
            quantity = 0
        unit_price = round(19.5 + random.random() * 980, 2)
        discount_rate = round((index % 7) * 0.03, 4)
        if index % 29 == 0:
            discount_rate = None
        source_sheet.append([
            index,
            f"SO-{2026 + index % 3}-{index:05d}",
            f"客户-{region}-{index:03d}",
            region,
            category,
            date(2025, 1, 1) + timedelta(days=index * 3),
            quantity,
            unit_price,
            discount_rate,
            f"销售-{index % 12 + 1:02d}",
            None,
            None,
            None,
            None,
            "含缺失值" if index % 29 == 0 else ("零数量" if quantity == 0 else "正常"),
        ])

    template = Workbook()
    detail = template.active
    detail.title = "销售明细"
    detail.append(headers)
    for row in range(2, ROWS + 2):
        detail.cell(row, 11, f"=G{row}*H{row}")
        detail.cell(row, 12, f"=K{row}*IFERROR(I{row},0)")
        detail.cell(row, 13, f"=K{row}-L{row}")
        detail.cell(row, 14, f'=IF(OR(G{row}=0,I{row}>0.15),"复核","正常")')
    summary = template.create_sheet("汇总")
    summary.append(["region", "total_net_amount", "order_count", "region_level", "owner_hint"])
    for row, region in enumerate(regions, start=2):
        summary.cell(row, 1, region)
        summary.cell(row, 2, f'=SUMIF(销售明细!D$2:D$201,A{row},销售明细!M$2:M$201)')
        summary.cell(row, 3, f'=COUNTIFS(销售明细!D$2:D$201,A{row},销售明细!N$2:N$201,"正常")')
        summary.cell(row, 4, f'=IFERROR(VLOOKUP(A{row},区域字典!$A:$B,2,FALSE),"未知")')
        summary.cell(row, 5, f'=IFERROR(XLOOKUP(A{row},区域字典!$A:$A,区域字典!$C:$C),"未配置")')
    lookup = template.create_sheet("区域字典")
    lookup.append(["region", "region_level", "owner"])
    for row, region in enumerate(regions, start=2):
        lookup.cell(row, 1, region)
        lookup.cell(row, 2, "核心区域" if row % 2 == 0 else "成长区域")
        lookup.cell(row, 3, f"区域负责人-{row - 1:02d}")

    thin = Side(style="thin", color="B7C3D0")
    header_fill = PatternFill("solid", fgColor="17324D")
    for worksheet in template.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(bottom=thin)
        for column in worksheet.columns:
            letter = column[0].column_letter
            worksheet.column_dimensions[letter].width = max(12, min(24, max(len(str(cell.value or "")) for cell in column) + 2))
    for row in range(2, ROWS + 2):
        detail.cell(row, 6).number_format = "yyyy-mm-dd"
        detail.cell(row, 8).number_format = '#,##0.00'
        detail.cell(row, 9).number_format = "0.00%"
        for column in (11, 12, 13):
            detail.cell(row, column).number_format = '#,##0.00'
    return source, template


def main() -> None:
    OUTPUT.mkdir(parents=True, exist_ok=True)
    source, template = build_workbook()
    source.save(OUTPUT / "complex_source_200x15.xlsx")
    template.save(OUTPUT / "complex_template_200x15.xlsx")
    print(f"generated {ROWS} rows x 15 columns in {OUTPUT}")


if __name__ == "__main__":
    main()
