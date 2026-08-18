import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from app.services.formula_service import function_names, inspect_formula_dependencies, preview_formula_results, python_aggregate, referenced_sheets, validate_formulas


class FormulaServiceTests(unittest.TestCase):
    def test_formula_functions_and_sheet_references(self):
        formula = '=IFERROR(VLOOKUP(A2,\'客户字典\'!$A:$B,2,FALSE),"未匹配")'
        self.assertEqual(function_names(formula), ["IFERROR", "VLOOKUP"])
        self.assertEqual(referenced_sheets(formula), ["客户字典"])

    def test_python_aggregate_matches_sumifs_countifs_shape(self):
        records = [{"region": "华东", "amount": 10, "risk": "正常"}, {"region": "华东", "amount": 5, "risk": "复核"}, {"region": "华南", "amount": 8, "risk": "正常"}]
        rows = python_aggregate(records, "region", "amount", {"risk": "正常"})
        self.assertEqual(rows, [{"group": "华东", "count": 1, "sum": 10.0}, {"group": "华南", "count": 1, "sum": 8.0}])

    def test_missing_sheet_is_reported(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "formula.xlsx"
            workbook = Workbook()
            workbook.active["A1"] = "=SUMIFS(不存在!A:A,A:A,1)"
            workbook.save(path)
            result = validate_formulas(str(path))
            self.assertFalse(result["valid"])
            self.assertEqual(result["issues"][0]["type"], "missing_sheet")

    def test_formula_preview_contains_formula_and_cached_value(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "formula.xlsx"
            workbook = Workbook()
            workbook.active["A1"] = 2
            workbook.active["B1"] = "=A1*3"
            workbook.save(path)
            result = preview_formula_results(str(path))
            self.assertEqual(result["formula_count"], 1)
            self.assertEqual(result["sheets"][0]["results"][0]["formula"], "=A1*3")
            self.assertIsNone(result["sheets"][0]["results"][0]["value"])

    def test_formula_dependencies_include_local_and_cross_sheet_cells(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "dependencies.xlsx"
            workbook = Workbook()
            workbook.active.title = "汇总"
            workbook.active["B2"] = "='明细'!C2+A2"
            workbook.create_sheet("明细")["C2"] = 5
            workbook.save(path)
            result = inspect_formula_dependencies(str(path))
            self.assertEqual(result["formula_count"], 1)
            self.assertEqual(result["dependencies"][0]["references"], [{"sheet": "明细", "cell": "C2"}, {"sheet": "汇总", "cell": "A2"}])
