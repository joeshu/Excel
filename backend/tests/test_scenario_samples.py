import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from app.services.data_quality import inspect_data_quality
from app.services.dag_engine import execute_dag
from app.services.data_reader import read_records
from app.services.formula_service import inspect_formulas, preview_formula_results, validate_formulas
from app.services.recalculation import recalculate
from app.services.template_parser import TemplateParser
from app.services.workflow_engine import WorkflowEngine


ROOT = Path(__file__).parents[2]
SCENARIO_DIR = ROOT / "sample_data" / "scenarios"


class ScenarioSampleTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        manifest_path = SCENARIO_DIR / "manifest.json"
        if not manifest_path.is_file():
            raise unittest.SkipTest("scenario manifest is not generated")
        cls.scenarios = json.loads(manifest_path.read_text(encoding="utf-8"))["scenarios"]

    def test_all_scenarios_parse_and_execute_formula_mode(self):
        self.assertEqual(len(self.scenarios), 4)
        for scenario in self.scenarios:
            with self.subTest(scenario=scenario["name"]):
                source_path = SCENARIO_DIR / scenario["source"]["path"]
                template_path = SCENARIO_DIR / scenario["template"]
                records = read_records(str(source_path))
                metadata = TemplateParser().parse(str(template_path))
                self.assertEqual(len(records), scenario["source"]["rows"])
                self.assertEqual(metadata["has_formula"], scenario["formula_count"] > 0)
                self.assertEqual(validate_formulas(str(template_path))["valid"], True)
                with tempfile.TemporaryDirectory() as directory:
                    output_path = Path(directory) / f"{scenario['name']}.xlsx"
                    engine = WorkflowEngine(str(template_path))
                    engine.execute_formula_mode(records, scenario["mapping"])
                    engine.save(str(output_path))
                    self.assertEqual(load_workbook(output_path, data_only=False).active.max_row, len(records) + 1)
                    self.assertEqual(validate_formulas(str(output_path))["valid"], True)
                    self.assertEqual(recalculate(str(output_path)).engine, "formula_only")

    def test_formula_and_no_formula_preview_paths(self):
        for scenario in self.scenarios:
            with self.subTest(scenario=scenario["name"]):
                template_path = SCENARIO_DIR / scenario["template"]
                inspection = inspect_formulas(str(template_path))
                preview = preview_formula_results(str(template_path), limit=1000)
                self.assertEqual(inspection["formula_count"], scenario["formula_count"])
                self.assertEqual(preview["formula_count"], scenario["formula_count"])

    def test_edge_case_source_quality_report_and_csv_path(self):
        scenario = next(item for item in self.scenarios if item["name"] == "quality_edge_cases")
        report = inspect_data_quality(str(SCENARIO_DIR / scenario["source"]["path"]))
        csv_records = read_records(str(SCENARIO_DIR / scenario["source"]["csv_path"]))
        self.assertFalse(report["valid"])
        self.assertGreater(report["issue_count"], 0)
        self.assertEqual(len(csv_records), scenario["source"]["rows"])

    def test_basic_no_formula_runs_mode_b_dag_end_to_end(self):
        scenario = next(item for item in self.scenarios if item["name"] == "basic_no_formula")
        source_path = SCENARIO_DIR / scenario["source"]["path"]
        template_path = SCENARIO_DIR / scenario["template"]
        node_json = {
            "nodes": [
                {"id": "source", "type": "data_source", "data": {"config": {"source_id": 1}}},
                {"id": "mapping", "type": "field_mapping", "data": {"config": {"mapping": scenario["mapping"]}}},
                {"id": "write", "type": "write_template", "data": {"config": {"mapping": scenario["mapping"]}}},
                {"id": "output", "type": "output_file"},
            ],
            "edges": [
                {"source": "source", "target": "mapping"},
                {"source": "mapping", "target": "write"},
                {"source": "write", "target": "output"},
            ],
        }
        with tempfile.TemporaryDirectory() as directory:
            output_path = Path(directory) / "basic_dag.xlsx"
            execute_dag(node_json, read_records(str(source_path)), str(template_path), str(output_path))
            workbook = load_workbook(output_path, data_only=False)
            self.assertEqual(workbook.active.max_row, scenario["source"]["rows"] + 1)
            self.assertEqual(workbook.active["B2"].value, "客户-001")
            self.assertEqual(validate_formulas(str(output_path))["formula_count"], 0)
            self.assertEqual(recalculate(str(output_path)).engine, "formula_only")


if __name__ == "__main__":
    unittest.main()
