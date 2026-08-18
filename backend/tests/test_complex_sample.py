import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from app.services.data_reader import read_records
from app.services.template_parser import TemplateParser
from app.services.workflow_engine import WorkflowEngine


class ComplexSampleTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        import sys
        sys.path.insert(0, str(Path(__file__).parents[2] / "tools"))
        from generate_complex_sample import build_workbook
        cls.source_workbook, cls.template_workbook = build_workbook()

    def test_200_rows_and_15_columns_are_read(self):
        with tempfile.TemporaryDirectory() as directory:
            source_path = Path(directory) / "source.xlsx"
            self.source_workbook.save(source_path)
            records = read_records(str(source_path))
            self.assertEqual(len(records), 200)
            self.assertEqual(len(records[0]), 15)
            self.assertEqual(records[0]["order_no"], "SO-2027-00001")

    def test_parser_and_engine_handle_complex_workbook(self):
        with tempfile.TemporaryDirectory() as directory:
            source_path = Path(directory) / "source.xlsx"
            template_path = Path(directory) / "template.xlsx"
            output_path = Path(directory) / "output.xlsx"
            self.source_workbook.save(source_path)
            self.template_workbook.save(template_path)
            metadata = TemplateParser().parse(str(template_path))
            self.assertEqual(metadata["sheet_count"], 2)
            self.assertTrue(metadata["has_formula"])
            records = read_records(str(source_path))
            mapping = {
                "销售明细!A": "record_id", "销售明细!B": "order_no", "销售明细!C": "customer_name",
                "销售明细!D": "region", "销售明细!E": "category", "销售明细!F": "order_date",
                "销售明细!G": "quantity", "销售明细!H": "unit_price", "销售明细!I": "discount_rate",
                "销售明细!J": "sales_owner", "销售明细!O": "remark",
            }
            engine = WorkflowEngine(str(template_path))
            engine.execute_formula_mode(records, mapping)
            engine.save(str(output_path))
            workbook = load_workbook(output_path, data_only=False)
            detail = workbook["销售明细"]
            summary = workbook["汇总"]
            self.assertEqual(detail.max_row, 201)
            self.assertEqual(detail["A201"].value, 200)
            self.assertEqual(detail["K201"].value, "=G201*H201")
            self.assertEqual(detail["N18"].value, '=IF(OR(G18=0,I18>0.15),"复核","正常")')
            self.assertEqual(summary["B2"].value, '=SUMIF(销售明细!D$2:D$201,A2,销售明细!M$2:M$201)')


if __name__ == "__main__":
    unittest.main()
