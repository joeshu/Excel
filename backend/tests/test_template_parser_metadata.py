import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from app.services.template_parser import TemplateParser


class TemplateParserMetadataTests(unittest.TestCase):
    def test_parses_hidden_columns_rows_freeze_filter_and_formula_references(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "template.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "明细"
            sheet.append(["地区", "金额", "合计"])
            sheet.append(["华东", 10, "=B2*2"])
            sheet.column_dimensions["B"].hidden = True
            sheet.row_dimensions[3].hidden = True
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = "A1:C2"
            sheet.merge_cells("A4:C4")
            workbook.create_sheet("字典").sheet_state = "hidden"
            workbook.save(path)
            metadata = TemplateParser().parse(str(path))
            detail = metadata["sheets"][0]
            amount = next(column for column in detail["columns"] if column["column"] == "B")
            total = next(column for column in detail["columns"] if column["column"] == "C")
            self.assertTrue(amount["hidden"])
            self.assertEqual(detail["hidden_rows"], [3])
            self.assertEqual(detail["freeze_panes"], "A2")
            self.assertEqual(detail["auto_filter"], "A1:C2")
            self.assertIn("A4:C4", detail["merged_ranges"])
            self.assertEqual(total["formula_references"], [{"sheet": "明细", "cell": "B2"}])
            self.assertTrue(metadata["sheets"][1]["hidden"])

    def test_identifies_detail_region_and_formula_risks(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "native-report.xlsx"
            workbook = Workbook()
            notice = workbook.active
            notice.title = "模版"
            notice.merge_cells("A1:C1")
            notice["A1"] = "通报"
            notice["A2"] = "=SUMIFS(明细!$C:$C,明细!$A:$A,\"南阳\")"
            detail = workbook.create_sheet("明细")
            detail.append(["日期", "单位", "金额"])
            detail.append(["说明", "说明", "说明"])
            detail.append(["字段", "字段", "字段"])
            detail.append(["20260731", "南阳", 10])
            workbook.save(path)
            metadata = TemplateParser().parse(str(path))
            notice_meta = metadata["sheets"][0]
            detail_meta = metadata["sheets"][1]
            self.assertEqual(notice_meta["role_candidates"][0]["role"], "notice")
            self.assertIsNone(notice_meta["data_start_row_candidate"])
            self.assertEqual(detail_meta["data_start_row_candidate"], 4)
            self.assertTrue(any(item["type"] == "whole_column_reference" for item in notice_meta["formula_risks"]))
            self.assertTrue(any(item["type"] == "hardcoded_dimension_value" for item in notice_meta["formula_risks"]))


if __name__ == "__main__":
    unittest.main()
