import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from app.services.workflow_engine import WorkflowEngine


class WorkflowEngineTests(unittest.TestCase):
    def create_template(self) -> tuple[Path, tempfile.TemporaryDirectory]:
        directory = tempfile.TemporaryDirectory()
        path = Path(directory.name) / "template.xlsx"
        workbook = Workbook()
        first = workbook.active
        first.title = "销售明细"
        first.append(["名称", "数量", "金额"])
        first.append([None, None, "=B2*10"])
        second = workbook.create_sheet("汇总")
        second.append(["名称"])
        second.append([None])
        workbook.save(path)
        return path, directory

    def test_fills_multiple_sheets_and_formula_rows(self):
        template, directory = self.create_template()
        self.addCleanup(directory.cleanup)
        engine = WorkflowEngine(str(template))
        engine.execute_formula_mode(
            [{"name": "A", "quantity": 2}, {"name": "B", "quantity": 3}],
            {"销售明细!A": "name", "销售明细!B": "quantity", "汇总!A": "name"},
        )
        output = Path(directory.name) / "output.xlsx"
        engine.save(str(output))
        workbook = load_workbook(output, data_only=False)
        self.assertEqual(workbook["销售明细"]["A2"].value, "A")
        self.assertEqual(workbook["销售明细"]["A3"].value, "B")
        self.assertEqual(workbook["销售明细"]["C3"].value, "=B3*10")
        self.assertEqual(workbook["汇总"]["A2"].value, "A")
        self.assertEqual(workbook["汇总"]["A3"].value, "B")

    def test_legacy_unqualified_mapping_only_writes_first_sheet(self):
        template, directory = self.create_template()
        self.addCleanup(directory.cleanup)
        engine = WorkflowEngine(str(template))
        engine.execute_formula_mode([{"name": "A"}], {"A": "name"})
        self.assertEqual(engine.workbook["销售明细"]["A2"].value, "A")
        self.assertIsNone(engine.workbook["汇总"]["A2"].value)

    def test_generated_rows_keep_template_style_and_height(self):
        template, directory = self.create_template()
        self.addCleanup(directory.cleanup)
        workbook = load_workbook(template)
        sheet = workbook["销售明细"]
        sheet.row_dimensions[2].height = 18
        sheet.row_dimensions[3].height = 28
        sheet["A3"].font = Font(name="Arial", bold=True, color="FF0000")
        sheet["B3"].fill = PatternFill("solid", fgColor="FFF2CC")
        workbook.save(template)

        engine = WorkflowEngine(str(template))
        engine.execute_formula_mode(
            [{"name": "A", "quantity": 2}, {"name": "B", "quantity": 3}, {"name": "C", "quantity": 4}],
            {"销售明细!A": "name", "销售明细!B": "quantity"},
        )
        output = Path(directory.name) / "styled-output.xlsx"
        engine.save(str(output))
        result = load_workbook(output)
        self.assertEqual(result["销售明细"].row_dimensions[4].height, 28)
        self.assertEqual(result["销售明细"]["A4"].font.color.rgb, "00FF0000")
        self.assertEqual(result["销售明细"]["B4"].fill.fgColor.rgb, "00FFF2CC")


if __name__ == "__main__":
    unittest.main()
