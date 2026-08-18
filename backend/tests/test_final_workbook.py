import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace

from openpyxl import Workbook, load_workbook

from app.services.final_workbook import append_final_sheets


class FinalWorkbookTests(unittest.TestCase):
    def test_appends_notice_source_config_and_quality_sheets(self):
        with tempfile.TemporaryDirectory() as directory:
            directory_path = Path(directory)
            source_path = directory_path / "source.xlsx"
            output_path = directory_path / "output.xlsx"
            source = Workbook()
            source.active.append(["id", "region", "amount"])
            source.active.append([1, "华东", 12])
            source.save(source_path)
            output = Workbook()
            output.active.title = "数据模板"
            output.active.append(["id", "region", "amount"])
            output.save(output_path)
            workflow = SimpleNamespace(name="示例工作流", mode="formula", column_mapping={"数据模板!A": "id"}, node_json={})
            template = SimpleNamespace(name="示例模板", version="1.0")
            data_source = SimpleNamespace(name="示例数据", source_type="upload", file_path=str(source_path))
            append_final_sheets(str(output_path), [{"id": 1, "region": "华东", "amount": 12}], workflow, template, data_source)
            workbook = load_workbook(output_path, data_only=False)
            self.assertEqual(workbook.sheetnames[:4], ["通报表", "数据模板", "基础数据", "工作流配置"])
            self.assertIn("数据质量报告", workbook.sheetnames)
            self.assertEqual(workbook["通报表"]["A6"].value, "id")
            self.assertEqual(workbook["基础数据"]["C2"].value, 12)
            self.assertEqual(workbook["工作流配置"]["B3"].value, "formula")
