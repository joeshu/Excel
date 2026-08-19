import shutil
import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import load_workbook

from app.services.template_native_engine import TemplateNativeEngine


SAMPLE_WORKBOOK = Path(__file__).parents[2] / ".monkeycode-tmp-files" / "f970ab10-模版+明细 - 副本.xlsx"


class TemplateNativeEngineTests(unittest.TestCase):
    def test_fills_detail_sheet_and_preserves_native_notice(self):
        if not SAMPLE_WORKBOOK.exists():
            self.skipTest("样例工作簿未上传")
        with TemporaryDirectory() as directory:
            template = Path(directory) / "template.xlsx"
            output = Path(directory) / "output.xlsx"
            shutil.copyfile(SAMPLE_WORKBOOK, template)
            records = [
                {"日期": "20260731", "单位": "南阳市邓州市分公司", "金额": 1.25},
                {"日期": "20260731", "单位": "南阳市镇平县分公司", "金额": 2.5},
            ]
            engine = TemplateNativeEngine(str(template), {"detail_sheet": "明细", "data_start_row": 4, "style_source_row": 4})
            engine.execute(records, {"明细!A": "日期", "明细!BA": "单位", "明细!W": "金额"})
            engine.save(str(output))

            workbook = load_workbook(output, data_only=False)
            notice = workbook["模版"]
            detail = workbook["明细"]
            self.assertEqual(workbook.sheetnames, ["模版", "明细"])
            self.assertEqual(detail["A4"].value, "20260731")
            self.assertEqual(detail["BA5"].value, "南阳市镇平县分公司")
            self.assertEqual(detail["W4"].value, 1.25)
            self.assertEqual(notice["A3"].value, "=明细!$A$5")
            self.assertEqual(notice["E5"].value, '=SUMIFS(明细!$W:$W,明细!$BA:$BA,$B:$B,明细!$B:$B,$A$3)')
            self.assertEqual([str(item) for item in notice.merged_cells.ranges], ["E2:J2", "G3:J3", "K2:L2", "C2:C4", "D2:D4", "C1:L1", "E3:F3", "K3:L3"])

    def test_copies_detail_style_when_records_exceed_template_rows(self):
        if not SAMPLE_WORKBOOK.exists():
            self.skipTest("样例工作簿未上传")
        with TemporaryDirectory() as directory:
            template = Path(directory) / "template.xlsx"
            output = Path(directory) / "output.xlsx"
            shutil.copyfile(SAMPLE_WORKBOOK, template)
            records = [{"日期": "20260731", "单位": "单位", "金额": index} for index in range(3500)]
            engine = TemplateNativeEngine(str(template), {"detail_sheet": "明细", "data_start_row": 4, "style_source_row": 4})
            engine.execute(records, {"明细!A": "日期", "明细!BA": "单位", "明细!W": "金额"})
            engine.save(str(output))
            workbook = load_workbook(output)
            detail = workbook["明细"]
            self.assertEqual(detail["W3503"].value, 3499)
            self.assertEqual(detail["W3503"].style_id, detail["W4"].style_id)


if __name__ == "__main__":
    unittest.main()
