import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.services.dag_engine import execute_dag, validate_dag


class DagEngineTests(unittest.TestCase):
    def test_validate_rejects_cycle(self):
        result = validate_dag({"nodes": [{"id": "a", "type": "data_source"}, {"id": "b", "type": "output_file"}], "edges": [{"source": "a", "target": "b"}, {"source": "b", "target": "a"}]})
        self.assertFalse(result["valid"])
        self.assertIn("循环", "；".join(result["issues"]))

    def test_validate_requires_connected_output(self):
        result = validate_dag({"nodes": [{"id": "source", "type": "data_source"}, {"id": "write", "type": "write_template"}, {"id": "output", "type": "output_file"}], "edges": [{"source": "source", "target": "write"}]})
        self.assertFalse(result["valid"])
        self.assertIn("输出文件节点", "；".join(result["issues"]))

    def test_execute_dag_filters_calculates_and_writes_template(self):
        with tempfile.TemporaryDirectory() as directory:
            template_path = Path(directory) / "template.xlsx"
            output_path = Path(directory) / "output.xlsx"
            workbook = Workbook()
            workbook.active.append(["id", "amount", "total"])
            workbook.save(template_path)
            node_json = {
                "nodes": [
                    {"id": "source", "type": "data_source"},
                    {"id": "formula", "type": "formula", "data": {"config": {"field": "total", "expression": "amount * 2"}}},
                    {"id": "condition", "type": "condition", "data": {"config": {"field": "total", "operator": "greater_than", "value": 10}}},
                    {"id": "write", "type": "write_template", "data": {"config": {"mapping": {"A": "id", "B": "amount", "C": "total"}}}},
                    {"id": "output", "type": "output_file"},
                ],
                "edges": [{"source": "source", "target": "formula"}, {"source": "formula", "target": "condition"}, {"source": "condition", "target": "write"}, {"source": "write", "target": "output"}],
            }
            execute_dag(node_json, [{"id": 1, "amount": 4}, {"id": 2, "amount": 8}], str(template_path), str(output_path))
            worksheet = load_workbook(output_path, data_only=False).active
            self.assertEqual(worksheet.max_row, 2)
            self.assertEqual(worksheet["A2"].value, 2)
            self.assertEqual(worksheet["C2"].value, 16)


if __name__ == "__main__":
    unittest.main()
