from copy import deepcopy

from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy import select
from sqlalchemy.orm import Session
from openpyxl import load_workbook

from app.database import get_db
from app.models.template import Template
from app.models.data_source import DataSource
from app.models.workflow import WorkflowDef
from app.models.mapping_snapshot import MappingSnapshot
from app.models.mapping_rule import MappingRule
from app.schemas.workflow import DagUpdate, MappingPreview, MappingRuleCreate, MappingSnapshotCreate, MappingUpdate, NoticeWorkflowConfigUpdate, WorkflowCreate, WorkflowWizardCreate
from app.services.dag_engine import validate_dag
from app.services.domain_metadata import field_signature
from app.services.mapping_service import data_fields, recommend_mapping, rules_from_workflow, template_columns, validate_mapping
from app.services.template_contract import validate_native_contract
from app.services.data_reader import read_records
from app.services.notice_calculation import calculate_notice
from app.services.notice_formula_profile import extract_notice_formula_profile
from app.services.template_parser import TemplateParser

router = APIRouter(prefix="/api/workflows", tags=["workflows"])


@router.post("", status_code=201)
def create_workflow(payload: WorkflowCreate, db: Session = Depends(get_db)):
    template = db.get(Template, payload.template_id)
    if not template:
        raise HTTPException(status_code=404, detail="模板不存在")
    notice_config = _default_notice_config(template) if payload.mode == "template_native" else {}
    if payload.mode == "template_native":
        _ensure_native_profile(template)
    workflow = WorkflowDef(template_id=payload.template_id, name=payload.name, mode=payload.mode, notice_config=notice_config)
    db.add(workflow)
    db.commit()
    db.refresh(workflow)
    return workflow


@router.post("/wizard", status_code=201)
def create_workflow_from_wizard(payload: WorkflowWizardCreate, db: Session = Depends(get_db)):
    source = db.get(DataSource, payload.data_source_id)
    template = db.get(Template, payload.template_id)
    if not source or not template:
        raise HTTPException(status_code=404, detail="数据源或模板不存在")
    notice_config = _default_notice_config(template) if payload.mode == "template_native" else {}
    if payload.mode == "template_native":
        _ensure_native_profile(template)
    workflow = WorkflowDef(template_id=payload.template_id, name=payload.name, mode=payload.mode, column_mapping=payload.column_mapping, node_json=payload.node_json, applicable_field_signature=source.field_signature, notice_config=notice_config)
    db.add(workflow)
    db.commit()
    db.refresh(workflow)
    return workflow


@router.get("")
def list_workflows(db: Session = Depends(get_db)):
    return db.scalars(select(WorkflowDef).order_by(WorkflowDef.updated_at.desc())).all()


def _default_notice_config(template: Template) -> dict:
    """Create editable defaults from the native notice layout without fixing business formulas."""
    try:
        workbook = load_workbook(template.file_path, read_only=True, data_only=False)
        notice_name = "模版" if "模版" in workbook.sheetnames else workbook.sheetnames[0]
        sheet = workbook[notice_name]
        rows = []
        for row_number in range(5, max(5, sheet.max_row)):
            key = sheet.cell(row_number, 2).value
            if key in {None, "", "合计"}:
                continue
            rows.append({"row": row_number, "key": str(key), "short_name": sheet.cell(row_number, 3).value, "target": sheet.cell(row_number, 4).value})
        metric_columns = {
            "daily": "E", "daily_rate": "F", "original": "G", "final": "H",
            "sequential_rate": "I", "rank": "J", "product_daily": "K", "product_monthly": "L",
        }
        metrics = {}
        for name, column in metric_columns.items():
            if name in {"daily_rate", "sequential_rate", "rank"}:
                metrics[name] = {"column": column, "kind": "derived", "source_metric": "daily"}
            else:
                metrics[name] = {"column": column, "source_field": "W", "aggregate": "sum", "dimension_field": "BA", "filters": []}
        total_row = next((row for row in range(5, sheet.max_row + 1) if sheet.cell(row, 2).value == "合计"), None)
        config = {"notice_sheet": notice_name, "dimensions": {"source_field": "BA", "rule_field": "AZ", "rule_value": "发展人"}, "rows": rows, "metrics": metrics, "totals": {}, "total_row": total_row, "execution_mode": "value"}
        if template.has_formula:
            return extract_notice_formula_profile(template.file_path, notice_name)
        return config
    except Exception:
        return {}


def _ensure_native_profile(template: Template) -> dict:
    metadata = TemplateParser().parse(template.file_path)
    profile = _recommended_profile_from_metadata(metadata)
    template.column_meta = {**(template.column_meta or {}), "native_profile": profile}
    return profile


def _recommended_profile_from_metadata(metadata: dict) -> dict:
    detail = next((item for item in metadata.get("sheets", []) if any(candidate["role"] == "detail" for candidate in item.get("role_candidates", []))), None)
    notice = next((item for item in metadata.get("sheets", []) if any(candidate["role"] == "notice" for candidate in item.get("role_candidates", []))), None)
    if not detail:
        return {}
    formula_headers = {item["header"] for item in detail.get("columns", []) if item.get("is_formula")}
    required = [item["header"] for item in detail.get("field_headers", []) if item.get("header") not in formula_headers]
    return {"notice_sheet": notice["title"] if notice else None, "detail_sheet": detail["title"], "data_start_row": detail.get("data_start_row_candidate") or 2, "style_source_row": detail.get("style_source_row_candidate") or detail.get("data_start_row_candidate") or 2, "data_end_rule": "last_nonempty_row", "field_contract": {"required": required, "required_source_fields": required, "derived_template_fields": sorted(formula_headers), "mapping": {item["column"]: item["header"] for item in detail.get("field_headers", []) if item["header"] not in formula_headers}}}


@router.get("/{workflow_id}/notice-config")
def get_notice_config(workflow_id: int, db: Session = Depends(get_db)):
    workflow = db.get(WorkflowDef, workflow_id)
    if not workflow:
        raise HTTPException(status_code=404, detail="工作流不存在")
    return {"workflow_id": workflow.id, "version": workflow.notice_config_version, "config": workflow.notice_config or {}, "history": workflow.notice_config_history or []}


@router.put("/{workflow_id}/notice-config")
def update_notice_config(workflow_id: int, payload: NoticeWorkflowConfigUpdate, db: Session = Depends(get_db)):
    workflow = db.get(WorkflowDef, workflow_id)
    if not workflow:
        raise HTTPException(status_code=404, detail="工作流不存在")
    config = payload.model_dump()
    dimensions = config.get("dimensions") or {}
    if workflow.mode == "template_native":
        source_field = dimensions.get("source_field")
        rule_field = dimensions.get("rule_field")
        if source_field != "BA":
            raise HTTPException(status_code=400, detail="模板原生通报的组织字段必须确认使用 BA（下沉区县名称）")
        if rule_field and rule_field != "AZ":
            raise HTTPException(status_code=400, detail="模板原生通报的规则字段必须使用 AZ（下沉规则）")
    if not config.get("rows"):
        raise HTTPException(status_code=400, detail="至少配置一行通报组织映射")
    if not config.get("metrics"):
        raise HTTPException(status_code=400, detail="至少配置一个通报指标")
    current_version = workflow.notice_config_version or 0
    history = list(workflow.notice_config_history or [])
    if workflow.notice_config:
        history.append({"version": current_version, "config": workflow.notice_config})
    workflow.notice_config_history = history
    workflow.notice_config = config
    workflow.notice_config_version = current_version + 1
    db.commit()
    db.refresh(workflow)
    return {"workflow_id": workflow.id, "version": workflow.notice_config_version, "config": workflow.notice_config}


@router.get("/{workflow_id}/notice-config/history")
def notice_config_history(workflow_id: int, db: Session = Depends(get_db)):
    workflow = db.get(WorkflowDef, workflow_id)
    if not workflow:
        raise HTTPException(status_code=404, detail="工作流不存在")
    return {"workflow_id": workflow.id, "current_version": workflow.notice_config_version, "history": workflow.notice_config_history or []}


@router.post("/{workflow_id}/notice-config/restore/{version}")
def restore_notice_config(workflow_id: int, version: int, db: Session = Depends(get_db)):
    workflow = db.get(WorkflowDef, workflow_id)
    if not workflow:
        raise HTTPException(status_code=404, detail="工作流不存在")
    previous = next((item for item in workflow.notice_config_history or [] if item.get("version") == version), None)
    if not previous:
        raise HTTPException(status_code=404, detail="历史通报配置版本不存在")
    current_version = workflow.notice_config_version or 0
    history = list(workflow.notice_config_history or [])
    history.append({"version": current_version, "config": workflow.notice_config or {}})
    workflow.notice_config_history = history
    workflow.notice_config = previous.get("config") or {}
    workflow.notice_config_version = current_version + 1
    db.commit()
    db.refresh(workflow)
    return {"workflow_id": workflow.id, "version": workflow.notice_config_version, "restored_from": version, "config": workflow.notice_config}


@router.post("/{workflow_id}/notice-config/preview")
def preview_notice_config(workflow_id: int, data_source_id: int, db: Session = Depends(get_db)):
    workflow = db.get(WorkflowDef, workflow_id)
    source = db.get(DataSource, data_source_id)
    if not workflow or not source:
        raise HTTPException(status_code=404, detail="工作流或数据源不存在")
    if not workflow.notice_config:
        raise HTTPException(status_code=400, detail="工作流尚未配置通报规则")
    try:
        result = calculate_notice(read_records(source.file_path), workflow.notice_config)
    except (OSError, ValueError) as error:
        raise HTTPException(status_code=400, detail=f"通报预览失败: {error}") from error
    return {"workflow_id": workflow.id, "data_source_id": source.id, "version": workflow.notice_config_version, **result}


@router.get("/{workflow_id}/contract/validate")
def validate_workflow_contract(workflow_id: int, data_source_id: int, db: Session = Depends(get_db)):
    workflow = db.get(WorkflowDef, workflow_id)
    source = db.get(DataSource, data_source_id)
    if not workflow or not source:
        raise HTTPException(status_code=404, detail="工作流或数据源不存在")
    if workflow.mode != "template_native":
        return {"valid": True, "required_fields": [], "missing_fields": [], "matched_fields": []}
    template = db.get(Template, workflow.template_id)
    profile = (template.column_meta or {}).get("native_profile", {}) if template else {}
    from app.models.template_workbook_profile import TemplateWorkbookProfile
    stored = db.scalar(select(TemplateWorkbookProfile).where(TemplateWorkbookProfile.template_id == template.id)) if template else None
    if stored:
        profile = stored.profile
    return validate_native_contract(profile, workflow.column_mapping, set(source.schema_ or {}))


@router.put("/{workflow_id}/mapping")
def update_mapping(workflow_id: int, payload: MappingUpdate, db: Session = Depends(get_db)):
    workflow = db.get(WorkflowDef, workflow_id)
    if not workflow:
        raise HTTPException(status_code=404, detail="工作流不存在")
    if workflow.mode not in {"formula", "manual"}:
        raise HTTPException(status_code=400, detail="当前模式不支持列映射")
    template = db.get(Template, workflow.template_id)
    sheets = template.column_meta.get("sheets", [])
    first_sheet = sheets[0]["title"] if sheets else ""
    normalized_mapping = {
        key if "!" in key else f"{first_sheet}!{key}": value
        for key, value in payload.column_mapping.items()
    }
    valid_columns = {
        f"{sheet['title']}!{item['column']}"
        for sheet in template.column_meta.get("sheets", [])
        for item in sheet.get("columns", [])
        if item.get("type") != "formula"
    }
    invalid_columns = sorted(set(normalized_mapping) - valid_columns)
    empty_fields = sorted(column for column in valid_columns if not normalized_mapping.get(column, "").strip())
    if invalid_columns:
        raise HTTPException(status_code=400, detail=f"存在不可映射的模板列: {', '.join(invalid_columns)}")
    if empty_fields:
        raise HTTPException(status_code=400, detail=f"以下模板列尚未配置数据源字段: {', '.join(empty_fields)}")
    workflow.column_mapping = normalized_mapping
    workflow.applicable_field_signature = field_signature(normalized_mapping.values())
    db.commit()
    db.refresh(workflow)
    return workflow


@router.post("/{workflow_id}/copy", status_code=201)
def copy_workflow(workflow_id: int, db: Session = Depends(get_db)):
    source = db.get(WorkflowDef, workflow_id)
    if not source:
        raise HTTPException(status_code=404, detail="工作流不存在")
    copied = WorkflowDef(
        template_id=source.template_id,
        name=f"{source.name} 副本",
        mode=source.mode,
        node_json=deepcopy(source.node_json or {}),
        column_mapping=deepcopy(source.column_mapping or {}),
        notice_config=deepcopy(source.notice_config or {}),
        notice_config_version=1,
        notice_config_history=[],
    )
    db.add(copied)
    db.commit()
    db.refresh(copied)
    return copied


@router.put("/{workflow_id}/dag")
def update_dag(workflow_id: int, payload: DagUpdate, db: Session = Depends(get_db)):
    workflow = db.get(WorkflowDef, workflow_id)
    if not workflow:
        raise HTTPException(status_code=404, detail="工作流不存在")
    if workflow.mode != "dag":
        raise HTTPException(status_code=400, detail="只有模式 B 支持流程节点")
    node_json = payload.model_dump()
    validation = validate_dag(node_json)
    if not validation["valid"]:
        raise HTTPException(status_code=400, detail="；".join(validation["issues"]))
    workflow.node_json = node_json
    db.commit()
    db.refresh(workflow)
    return {"workflow": workflow, "validation": validation}


@router.get("/{workflow_id}/dag/validate")
def validate_workflow_dag(workflow_id: int, db: Session = Depends(get_db)):
    workflow = db.get(WorkflowDef, workflow_id)
    if not workflow:
        raise HTTPException(status_code=404, detail="工作流不存在")
    return validate_dag(workflow.node_json or {})


@router.get("/{workflow_id}/mapping-rules")
def workflow_mapping_rules(workflow_id: int, db: Session = Depends(get_db)):
    workflow = db.get(WorkflowDef, workflow_id)
    if not workflow:
        raise HTTPException(status_code=404, detail="工作流不存在")
    template = db.get(Template, workflow.template_id)
    if not template:
        raise HTTPException(status_code=404, detail="模板不存在")
    rules = rules_from_workflow(workflow)
    return {
        "workflow_id": workflow.id,
        "mode": workflow.mode,
        "template_id": template.id,
        "template_version": template.version,
        "columns": template_columns(template.column_meta or {}),
        "rules": rules,
    }


@router.post("/{workflow_id}/mapping-rules", status_code=201)
def create_mapping_rule(workflow_id: int, payload: MappingRuleCreate, db: Session = Depends(get_db)):
    workflow = db.get(WorkflowDef, workflow_id)
    source = db.get(DataSource, payload.data_source_id)
    template = db.get(Template, workflow.template_id) if workflow else None
    if not workflow or not source or not template:
        raise HTTPException(status_code=404, detail="工作流、模板或数据源不存在")
    validation = validate_mapping(template_columns(template.column_meta or {}), data_fields(source.schema_ or {}), payload.rules)
    if not validation["valid"]:
        raise HTTPException(status_code=400, detail={"message": "映射预检未通过", "validation": validation})
    latest = db.scalar(select(MappingRule.version).where(MappingRule.workflow_id == workflow.id).order_by(MappingRule.version.desc()).limit(1))
    rule = MappingRule(
        workflow_id=workflow.id,
        template_id=template.id,
        template_version=template.version,
        data_field_signature=source.field_signature,
        version=(latest or 0) + 1,
        rules=payload.rules,
        dependency_order=validation["dependency_order"],
        validation_result=validation,
    )
    db.add(rule)
    db.commit()
    db.refresh(rule)
    return rule


@router.get("/{workflow_id}/mapping-rules/versions")
def list_mapping_rule_versions(workflow_id: int, db: Session = Depends(get_db)):
    workflow = db.get(WorkflowDef, workflow_id)
    if not workflow:
        raise HTTPException(status_code=404, detail="工作流不存在")
    return db.scalars(select(MappingRule).where(MappingRule.workflow_id == workflow_id).order_by(MappingRule.version.desc())).all()


@router.get("/{workflow_id}/mapping-rules/versions/{version}")
def get_mapping_rule_version(workflow_id: int, version: int, db: Session = Depends(get_db)):
    rule = db.scalar(select(MappingRule).where(MappingRule.workflow_id == workflow_id, MappingRule.version == version))
    if not rule:
        raise HTTPException(status_code=404, detail="映射规则版本不存在")
    return rule


@router.post("/mapping/preview")
def preview_mapping(payload: MappingPreview, db: Session = Depends(get_db)):
    template = db.get(Template, payload.template_id)
    source = db.get(DataSource, payload.data_source_id)
    if not template or not source:
        raise HTTPException(status_code=404, detail="模板或数据源不存在")
    columns = template_columns(template.column_meta or {})
    fields = data_fields(source.schema_ or {})
    recommendations = recommend_mapping(columns, fields)
    validation = validate_mapping(columns, fields, payload.rules)
    return {"template_id": template.id, "template_version": template.version, "data_source_id": source.id, "data_field_signature": source.field_signature, "columns": columns, "fields": fields, "recommendations": recommendations, "validation": validation}


@router.post("/mapping/validate")
def validate_mapping_config(payload: MappingPreview, db: Session = Depends(get_db)):
    template = db.get(Template, payload.template_id)
    source = db.get(DataSource, payload.data_source_id)
    if not template or not source:
        raise HTTPException(status_code=404, detail="模板或数据源不存在")
    return validate_mapping(template_columns(template.column_meta or {}), data_fields(source.schema_ or {}), payload.rules)


@router.post("/{workflow_id}/mapping-snapshots", status_code=201)
def create_mapping_snapshot(workflow_id: int, payload: MappingSnapshotCreate, db: Session = Depends(get_db)):
    workflow = db.get(WorkflowDef, workflow_id)
    source = db.get(DataSource, payload.data_source_id)
    template = db.get(Template, workflow.template_id) if workflow else None
    if not workflow or not source or not template:
        raise HTTPException(status_code=404, detail="工作流、模板或数据源不存在")
    validation = validate_mapping(template_columns(template.column_meta or {}), data_fields(source.schema_ or {}), payload.rules)
    if not validation["valid"]:
        raise HTTPException(status_code=400, detail={"message": "映射预检未通过", "validation": validation})
    latest_rule_version = db.scalar(select(MappingRule.version).where(MappingRule.workflow_id == workflow.id).order_by(MappingRule.version.desc()).limit(1))
    snapshot = MappingSnapshot(
        workflow_id=workflow.id,
        template_id=template.id,
        data_source_id=source.id,
        template_version=template.version,
        data_field_signature=source.field_signature,
        rule_version=latest_rule_version or 1,
        rules=payload.rules,
        dependency_order=validation["dependency_order"],
        validation_result=validation,
    )
    db.add(snapshot)
    db.commit()
    db.refresh(snapshot)
    return snapshot


@router.get("/{workflow_id}/mapping-snapshots")
def list_mapping_snapshots(workflow_id: int, db: Session = Depends(get_db)):
    workflow = db.get(WorkflowDef, workflow_id)
    if not workflow:
        raise HTTPException(status_code=404, detail="工作流不存在")
    snapshots = db.scalars(select(MappingSnapshot).where(MappingSnapshot.workflow_id == workflow_id).order_by(MappingSnapshot.created_at.desc())).all()
    return snapshots
