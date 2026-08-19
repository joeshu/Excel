from pathlib import Path
from uuid import uuid4

from fastapi import APIRouter, Body, Depends, File, Form, HTTPException, UploadFile, status
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.config import settings
from app.database import get_db
from app.models.template import Template
from app.models.template_workbook_profile import TemplateWorkbookProfile
from app.services.template_parser import TemplateParser, ensure_xlsx
from app.services.mapping_service import template_columns as normalize_template_columns

router = APIRouter(prefix="/api/templates", tags=["templates"])


def _template_or_404(template_id: int, db: Session) -> Template:
    template = db.get(Template, template_id)
    if not template:
        raise HTTPException(status_code=404, detail="模板不存在")
    return template


@router.post("/upload", status_code=status.HTTP_201_CREATED)
async def upload_template(file: UploadFile = File(...), parent_template_id: int | None = Form(None), version: str = Form("1.0"), db: Session = Depends(get_db)):
    try:
        ensure_xlsx(file.filename or "")
    except ValueError as error:
        raise HTTPException(status_code=400, detail=str(error)) from error
    Path(settings.upload_dir).mkdir(parents=True, exist_ok=True)
    path = Path(settings.upload_dir) / f"{uuid4().hex}_{Path(file.filename).name}"
    path.write_bytes(await file.read())
    try:
        metadata = TemplateParser().parse(str(path))
    except Exception as error:
        raise HTTPException(status_code=400, detail=f"模板解析失败: {error}") from error
    parent = db.get(Template, parent_template_id) if parent_template_id else None
    if parent_template_id and not parent:
        raise HTTPException(status_code=404, detail="父模板不存在")
    template = Template(name=parent.name if parent else Path(file.filename).stem, file_path=str(path), has_formula=metadata["has_formula"], column_meta=metadata, sheet_count=metadata["sheet_count"], version=version.strip() or "1.0", parent_template_id=parent_template_id)
    db.add(template)
    db.commit()
    db.refresh(template)
    return template


@router.get("")
def list_templates(db: Session = Depends(get_db)):
    return db.scalars(select(Template).order_by(Template.created_at.desc())).all()


@router.get("/{template_id}/versions")
def template_versions(template_id: int, db: Session = Depends(get_db)):
    template = db.get(Template, template_id)
    if not template:
        raise HTTPException(status_code=404, detail="模板不存在")
    root_id = template.parent_template_id or template.id
    versions = db.scalars(select(Template).where((Template.id == root_id) | (Template.parent_template_id == root_id)).order_by(Template.created_at.desc())).all()
    return {"template_id": template_id, "root_template_id": root_id, "versions": versions}


@router.get("/{template_id}/columns")
def template_columns(template_id: int, db: Session = Depends(get_db)):
    template = db.get(Template, template_id)
    if not template:
        raise HTTPException(status_code=404, detail="模板不存在")
    return template.column_meta


@router.get("/{template_id}/mapping-schema")
def template_mapping_schema(template_id: int, db: Session = Depends(get_db)):
    template = _template_or_404(template_id, db)
    return {
        "template_id": template.id,
        "template_version": template.version,
        "name": template.name,
        "columns": normalize_template_columns(template.column_meta or {}),
    }


@router.post("/{template_id}/analyze-workbook")
def analyze_workbook(template_id: int, db: Session = Depends(get_db)):
    template = _template_or_404(template_id, db)
    metadata = TemplateParser().parse(template.file_path)
    return {
        "template_id": template.id,
        "template_version": template.version,
        "sheets": metadata.get("sheets", []),
        "recommended_profile": _recommended_native_profile(metadata),
    }


@router.get("/{template_id}/workbook-profile")
def get_workbook_profile(template_id: int, db: Session = Depends(get_db)):
    template = _template_or_404(template_id, db)
    stored = db.scalar(select(TemplateWorkbookProfile).where(TemplateWorkbookProfile.template_id == template.id))
    profile = stored.profile if stored else (template.column_meta or {}).get("native_profile", {})
    return {"template_id": template.id, "template_version": template.version, "profile": profile, "updated_at": stored.updated_at if stored else None}


@router.put("/{template_id}/workbook-profile")
def save_workbook_profile(template_id: int, profile: dict = Body(...), db: Session = Depends(get_db)):
    template = _template_or_404(template_id, db)
    stored = db.scalar(select(TemplateWorkbookProfile).where(TemplateWorkbookProfile.template_id == template.id))
    if stored:
        stored.profile = profile
    else:
        stored = TemplateWorkbookProfile(template_id=template.id, profile=profile, validation_result={})
        db.add(stored)
    db.commit()
    db.refresh(template)
    return {"template_id": template.id, "template_version": template.version, "profile": profile, "updated_at": stored.updated_at}


@router.post("/{template_id}/workbook-profile/validate")
def validate_workbook_profile(template_id: int, profile: dict | None = Body(None), db: Session = Depends(get_db)):
    template = _template_or_404(template_id, db)
    metadata = TemplateParser().parse(template.file_path)
    stored = db.scalar(select(TemplateWorkbookProfile).where(TemplateWorkbookProfile.template_id == template.id))
    selected = profile if profile is not None else (stored.profile if stored else (template.column_meta or {}).get("native_profile", {}))
    sheets = {item["title"]: item for item in metadata.get("sheets", [])}
    errors = []
    warnings = []
    detail_title = selected.get("detail_sheet")
    if not detail_title:
        errors.append({"type": "missing_detail_sheet", "message": "未配置明细 Sheet"})
    elif detail_title not in sheets:
        errors.append({"type": "detail_sheet_not_found", "message": f"明细 Sheet 不存在: {detail_title}"})
    if selected.get("data_start_row") is not None and int(selected["data_start_row"]) < 1:
        errors.append({"type": "invalid_data_start_row", "message": "数据起始行必须大于 0"})
    for sheet in metadata.get("sheets", []):
        for risk in sheet.get("formula_risks", []):
            warnings.append({"sheet": sheet["title"], **risk})
    result = {"valid": not errors, "errors": errors, "warnings": warnings, "profile": selected}
    if stored and profile is None:
        stored.validation_result = result
        db.commit()
    return result


def _recommended_native_profile(metadata: dict) -> dict:
    detail = next((item for item in metadata.get("sheets", []) if any(candidate["role"] == "detail" for candidate in item.get("role_candidates", []))), None)
    notice = next((item for item in metadata.get("sheets", []) if any(candidate["role"] == "notice" for candidate in item.get("role_candidates", []))), None)
    if not detail:
        return {}
    formula_headers = {item["header"] for item in detail.get("columns", []) if item.get("is_formula")}
    required_fields = [item["header"] for item in detail.get("field_headers", []) if item.get("header") not in formula_headers]
    derived_fields = [item["header"] for item in detail.get("field_headers", []) if item.get("header") in formula_headers]
    return {
        "notice_sheet": notice["title"] if notice else None,
        "detail_sheet": detail["title"],
        "data_start_row": detail.get("data_start_row_candidate") or 2,
        "style_source_row": detail.get("style_source_row_candidate") or detail.get("data_start_row_candidate") or 2,
        "data_end_rule": "last_nonempty_row",
        "field_contract": {"required": required_fields, "required_source_fields": required_fields, "derived_template_fields": derived_fields, "mapping": {item["column"]: item["header"] for item in detail.get("field_headers", []) if item["header"] not in formula_headers}},
    }


@router.get("/{template_id}/preview")
def template_preview(template_id: int, limit: int = 20, db: Session = Depends(get_db)):
    template = db.get(Template, template_id)
    if not template:
        raise HTTPException(status_code=404, detail="模板不存在")
    workbook = TemplateParser().parse(template.file_path)
    from openpyxl import load_workbook
    source = load_workbook(template.file_path, read_only=True, data_only=False)
    sheets = []
    for worksheet in source.worksheets:
        rows = list(worksheet.iter_rows(min_row=1, max_row=max(1, min(limit, worksheet.max_row)), values_only=True))
        sheets.append({"title": worksheet.title, "rows": rows})
    return {"metadata": workbook, "sheets": sheets}
