from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.formula.translate import Translator

from app.services.notice_calculation import calculate_notice
from app.services.notice_formula import emit_notice_formulas


class TemplateNativeEngine:
    """Fills a configured detail sheet while preserving the native report sheet."""

    def __init__(self, template_path: str, profile: dict | None = None):
        self.workbook = load_workbook(template_path, read_only=False, data_only=False)
        self.profile = profile or {}

    def execute(self, records: list[dict], column_mapping: dict[str, str]) -> "TemplateNativeEngine":
        worksheet = self._detail_sheet()
        start_row = int(self.profile.get("data_start_row") or self._detect_data_start_row(worksheet))
        style_source_row = int(self.profile.get("style_source_row") or start_row)
        old_end_row = int(self.profile.get("data_end_row") or max(start_row, worksheet.max_row))
        configured_mapping = column_mapping or (self.profile.get("field_contract", {}) or {}).get("mapping", {})
        mappings = self._normalise_mapping(worksheet.title, configured_mapping)
        formula_columns = self._formula_columns(worksheet, start_row, old_end_row)

        for row_number in range(start_row, max(old_end_row, start_row + len(records) - 1) + 1):
            if row_number > old_end_row:
                self._copy_row_format(worksheet, row_number, style_source_row)
            for column in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_number, column=column)
                if isinstance(cell, MergedCell) or column in formula_columns:
                    continue
                cell.value = None

        for offset, record in enumerate(records):
            row_number = start_row + offset
            if row_number > old_end_row:
                self._copy_row_format(worksheet, row_number, style_source_row)
            for column, field in mappings.items():
                cell = worksheet.cell(row=row_number, column=column)
                if isinstance(cell, MergedCell) or column in formula_columns:
                    continue
                self._write_preserving_style(cell, record.get(field))
            self._fill_formulas(worksheet, row_number, formula_columns)

        self._update_filter(worksheet, start_row, len(records))
        self.workbook.calculation.fullCalcOnLoad = True
        self.workbook.calculation.forceFullCalc = True
        self.workbook.calculation.calcMode = "auto"
        return self

    def execute_notice(self, records: list[dict], config: dict) -> dict:
        """Calculate and write configured notice cells while preserving workbook layout."""
        notice_sheet_name = config.get("notice_sheet") or self.profile.get("notice_sheet") or "模版"
        if notice_sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"模板通报 Sheet 不存在: {notice_sheet_name}")
        runtime_config = dict(config)
        runtime_config["dimensions"] = dict(config.get("dimensions") or {})
        for metric in (runtime_config.get("metrics") or {}).values():
            date_config = metric.get("date") or {}
            if date_config.get("value_ref"):
                date_config = dict(date_config)
                date_config["value"] = self.workbook[notice_sheet_name][date_config["value_ref"]].value
                metric["date"] = date_config
        result = calculate_notice(records, runtime_config)
        worksheet = self.workbook[notice_sheet_name]
        for row in config.get("rows") or []:
            row_number = int(row["row"])
            row_values = result["values"].get(str(row_number), {})
            for metric_name, metric in (config.get("metrics") or {}).items():
                column = metric.get("column")
                if not column or metric_name not in row_values:
                    continue
                worksheet[f"{column}{row_number}"] = row_values[metric_name]
        total_row = config.get("total_row")
        if total_row:
            for metric_name, metric in (config.get("metrics") or {}).items():
                column = metric.get("column")
                if column and metric_name in result["totals"]:
                    worksheet[f"{column}{int(total_row)}"] = result["totals"][metric_name]
        if config.get("execution_mode") == "formula":
            formulas = emit_notice_formulas(runtime_config, self._detail_sheet().title)
            for row_number, row_formulas in formulas.items():
                for metric_name, formula in row_formulas.items():
                    column = (config.get("metrics") or {}).get(metric_name, {}).get("column")
                    if column:
                        worksheet[f"{column}{row_number}"] = formula
        return result

    def save(self, output_path: str) -> str:
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(output_path)
        return output_path

    def _detail_sheet(self):
        configured = self.profile.get("detail_sheet") or self.profile.get("sheet")
        if configured:
            if configured not in self.workbook.sheetnames:
                raise ValueError(f"模板明细 Sheet 不存在: {configured}")
            return self.workbook[configured]
        for worksheet in self.workbook.worksheets:
            if worksheet.title in {"明细", "明细数据", "detail", "data"} or worksheet.max_row > 100:
                return worksheet
        raise ValueError("未识别到模板明细 Sheet，请配置 detail_sheet")

    @staticmethod
    def _detect_data_start_row(worksheet) -> int:
        for row_number in range(1, min(worksheet.max_row, 20) + 1):
            values = [cell.value for cell in worksheet[row_number]]
            next_values = [cell.value for cell in worksheet[row_number + 1]] if row_number < worksheet.max_row else []
            if sum(value is not None for value in values) >= 2 and sum(value is not None for value in next_values) >= 2:
                if any(isinstance(value, (int, float)) for value in next_values):
                    return row_number + 1
        return 2

    @staticmethod
    def _normalise_mapping(sheet_title: str, mapping: dict[str, str]) -> dict[int, str]:
        result = {}
        for key, field in mapping.items():
            if key.startswith(f"{sheet_title}!"):
                key = key.split("!", 1)[1]
            if "!" in key or key.startswith("__"):
                continue
            try:
                from openpyxl.utils.cell import column_index_from_string
                result[column_index_from_string(key)] = field
            except ValueError:
                continue
        return result

    @staticmethod
    def _formula_columns(worksheet, start_row: int, end_row: int) -> set[int]:
        columns = set()
        for row in range(start_row, min(end_row, worksheet.max_row) + 1):
            for cell in worksheet[row]:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    columns.add(cell.column)
        return columns

    @staticmethod
    def _copy_row_format(worksheet, row_number: int, source_row: int) -> None:
        source_dimension = worksheet.row_dimensions[source_row]
        target_dimension = worksheet.row_dimensions[row_number]
        target_dimension.height = source_dimension.height
        target_dimension.hidden = source_dimension.hidden
        target_dimension.outlineLevel = source_dimension.outlineLevel
        target_dimension.collapsed = source_dimension.collapsed
        for column in range(1, worksheet.max_column + 1):
            source = worksheet.cell(row=source_row, column=column)
            target = worksheet.cell(row=row_number, column=column)
            if isinstance(target, MergedCell):
                continue
            target._style = copy(source._style)
            target.number_format = source.number_format
            target.alignment = copy(source.alignment)
            target.protection = copy(source.protection)

    @staticmethod
    def _write_preserving_style(cell, value) -> None:
        style = copy(cell._style)
        cell.value = value
        cell._style = style

    @staticmethod
    def _fill_formulas(worksheet, row_number: int, formula_columns: set[int]) -> None:
        for column in formula_columns:
            source = next((worksheet.cell(row=row, column=column) for row in range(1, row_number) if isinstance(worksheet.cell(row=row, column=column).value, str) and worksheet.cell(row=row, column=column).value.startswith("=")), None)
            target = worksheet.cell(row=row_number, column=column)
            if source is not None and target.value is None:
                target.value = Translator(source.value, origin=source.coordinate).translate_formula(target.coordinate)

    @staticmethod
    def _update_filter(worksheet, start_row: int, record_count: int) -> None:
        if worksheet.auto_filter.ref:
            start, end = worksheet.auto_filter.ref.split(":", 1)
            end_column = "".join(char for char in end if char.isalpha())
            worksheet.auto_filter.ref = f"{start}:{end_column}{max(start_row, start_row + record_count - 1)}"
