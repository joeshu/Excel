import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

SUPPORTED_FUNCTIONS = {"VLOOKUP", "XLOOKUP", "SUMIF", "SUMIFS", "COUNTIF", "COUNTIFS", "IF", "IFERROR", "AND", "OR"}
ERROR_VALUES = {"#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NUM!", "#NULL!"}
SHEET_REFERENCE = re.compile(r"(?:'([^']+)'|([^\s!+\-*/(),]+))!")
FUNCTION_NAME = re.compile(r"\b([A-Z][A-Z0-9_.]*)\s*\(")
CELL_REFERENCE = re.compile(r"(?:(?:'([^']+)'|([A-Za-z0-9_][^!+\-*/(), ]*))!)?\$?([A-Z]{1,3})\$?(\d+)")


def inspect_formulas(file_path: str) -> dict:
    workbook = load_workbook(file_path, read_only=False, data_only=False)
    sheets = []
    for worksheet in workbook.worksheets:
        formulas = []
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append({"cell": cell.coordinate, "formula": cell.value, "functions": function_names(cell.value)})
        sheets.append({"title": worksheet.title, "formula_count": len(formulas), "formulas": formulas})
    return {"sheets": sheets, "formula_count": sum(sheet["formula_count"] for sheet in sheets)}


def validate_formulas(file_path: str) -> dict:
    workbook = load_workbook(file_path, read_only=False, data_only=False)
    known_sheets = set(workbook.sheetnames)
    issues = []
    formula_count = 0
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str) or not cell.value.startswith("="):
                    continue
                formula_count += 1
                functions = function_names(cell.value)
                unsupported = sorted(set(functions) - SUPPORTED_FUNCTIONS)
                for function in unsupported:
                    issues.append({"sheet": worksheet.title, "cell": cell.coordinate, "type": "unsupported_function", "message": f"函数 {function} 尚未纳入静态校验"})
                for referenced_sheet in referenced_sheets(cell.value):
                    if referenced_sheet not in known_sheets:
                        issues.append({"sheet": worksheet.title, "cell": cell.coordinate, "type": "missing_sheet", "message": f"引用的 Sheet 不存在: {referenced_sheet}"})
    return {"valid": not issues, "formula_count": formula_count, "issues": issues}


def find_cached_errors(file_path: str) -> dict:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    errors = []
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value in ERROR_VALUES:
                    errors.append({"sheet": worksheet.title, "cell": cell.coordinate, "value": cell.value})
    return {"valid": not errors, "errors": errors}


def preview_formula_results(file_path: str, limit: int = 100) -> dict:
    formulas_workbook = load_workbook(file_path, read_only=True, data_only=False)
    values_workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheets = []
    formula_count = 0
    for formulas_sheet, values_sheet in zip(formulas_workbook.worksheets, values_workbook.worksheets):
        results = []
        max_row = min(formulas_sheet.max_row, max(1, limit))
        for row in formulas_sheet.iter_rows(min_row=1, max_row=max_row):
            for formula_cell in row:
                if not isinstance(formula_cell.value, str) or not formula_cell.value.startswith("="):
                    continue
                value = values_sheet[formula_cell.coordinate].value
                results.append({
                    "cell": formula_cell.coordinate,
                    "formula": formula_cell.value,
                    "value": value,
                    "error": value if value in ERROR_VALUES else None,
                })
        formula_count += len(results)
        sheets.append({"title": formulas_sheet.title, "formula_count": len(results), "results": results})
    return {"formula_count": formula_count, "sheets": sheets}


def inspect_formula_dependencies(file_path: str) -> dict:
    workbook = load_workbook(file_path, read_only=True, data_only=False)
    dependencies = []
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str) or not cell.value.startswith("="):
                    continue
                references = []
                for quoted_sheet, plain_sheet, column, row_number in CELL_REFERENCE.findall(cell.value):
                    references.append({
                        "sheet": quoted_sheet or plain_sheet or worksheet.title,
                        "cell": f"{column.upper()}{row_number}",
                    })
                dependencies.append({"sheet": worksheet.title, "cell": cell.coordinate, "formula": cell.value, "references": references})
    return {"formula_count": len(dependencies), "dependencies": dependencies}


def python_aggregate(records: list[dict], group_field: str, value_field: str, filters: dict[str, object] | None = None) -> list[dict]:
    filters = filters or {}
    grouped: dict[object, dict[str, object]] = defaultdict(lambda: {"count": 0, "sum": 0.0})
    for record in records:
        if any(record.get(key) != expected for key, expected in filters.items()):
            continue
        group = record.get(group_field)
        value = record.get(value_field)
        if group is None:
            continue
        try:
            numeric_value = float(value or 0)
        except (TypeError, ValueError):
            numeric_value = 0.0
        grouped[group]["count"] += 1
        grouped[group]["sum"] += numeric_value
    return [{"group": group, "count": round(values["count"], 0), "sum": round(values["sum"], 2)} for group, values in grouped.items()]


def function_names(formula: str) -> list[str]:
    return sorted(set(match.upper() for match in FUNCTION_NAME.findall(formula.upper())))


def referenced_sheets(formula: str) -> list[str]:
    return sorted(set(quoted or plain for quoted, plain in SHEET_REFERENCE.findall(formula)))
