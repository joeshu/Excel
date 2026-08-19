from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook


_SUMIFS_RE = re.compile(
    r"SUMIFS\((?P<sum_sheet>[^!]+)!\$(?P<sum_col>[A-Z]+):\$(?P=sum_col),"
    r"(?P<criteria>.+)\)$",
    re.IGNORECASE,
)


def extract_notice_formula_profile(path: str, notice_sheet: str = "模版") -> dict:
    """Convert a formula workbook into editable, layout-independent rule metadata."""
    workbook = load_workbook(Path(path), read_only=True, data_only=False)
    if notice_sheet not in workbook.sheetnames:
        raise ValueError(f"公式模板缺少通报 Sheet: {notice_sheet}")
    sheet = workbook[notice_sheet]
    rows = []
    for row_number in range(5, sheet.max_row + 1):
        name = sheet.cell(row_number, 2).value
        if name not in (None, "", "合计"):
            rows.append({"row": row_number, "key": str(name), "short_name": sheet.cell(row_number, 3).value, "target": sheet.cell(row_number, 4).value})
    total_row = next((row for row in range(5, sheet.max_row + 1) if sheet.cell(row, 2).value == "合计"), None)
    first_row = rows[0]["row"] if rows else 5
    formulas = {column: sheet[f"{column}{first_row}"].value for column in "EFGHIJKL"}
    detail_sheet = _detail_sheet(formulas, workbook.sheetnames)
    detail_headers = _detail_headers(workbook[detail_sheet])
    rule_field, rule_value = _rule_from_formulas(formulas, detail_headers)
    profile = {
        "notice_sheet": notice_sheet,
        "detail_sheet": detail_sheet,
        "dimensions": {"source_field": detail_headers.get("BA", "BA"), "source_column": "BA", "rule_field": rule_field, "rule_value": rule_value, "date_field": detail_headers.get("B", "B"), "date_column": "B", "date_cell": f"'{notice_sheet}'!$A$3"},
        "rows": rows,
        "metrics": {},
        "totals": {},
        "total_row": total_row,
        "execution_mode": "value",
        "formula_source": {"cells": formulas, "rank_ranges": _rank_ranges(sheet, rows)},
    }
    profile["metrics"] = _metrics_from_formulas(formulas, profile["formula_source"]["rank_ranges"], detail_headers)
    return profile


def _detail_sheet(formulas: dict, sheet_names: list[str]) -> str:
    for formula in formulas.values():
        if isinstance(formula, str) and "!" in formula:
            name = formula.split("!", 1)[0].replace("'", "")
            if name in sheet_names:
                return name
    return "明细" if "明细" in sheet_names else sheet_names[-1]


def _detail_headers(sheet) -> dict[str, str]:
    return {cell.column_letter: str(cell.value).strip() for cell in sheet[3] if cell.value not in {None, ""}}


def _rank_ranges(sheet, rows: list[dict]) -> list[dict]:
    result = []
    for item in rows:
        formula = sheet[f"J{item['row']}"].value
        match = re.search(r"\$I\$(\d+):\$I\$(\d+)", str(formula or ""))
        if match:
            result.append({"row": item["row"], "first": int(match.group(1)), "last": int(match.group(2))})
    return result


def _rule_from_formulas(formulas: dict, headers: dict[str, str]) -> tuple[str | None, str | None]:
    """Extract a shared SUMIFS rule while leaving unrelated metric filters local."""
    for formula in formulas.values():
        text = str(formula or "")
        for field_column, value in re.findall(r"!\$([A-Z]+):\$\1,\s*\"([^\"]+)\"", text, re.IGNORECASE):
            if field_column not in {"B", "C", "BA"}:
                return field_column, value
    return None, None


def _metrics_from_formulas(formulas: dict, rank_ranges: list[dict], headers: dict[str, str]) -> dict:
    daily = _sum_metric(formulas.get("E"), "E", headers, date_scope="day")
    daily["date"]["value_ref"] = "A3"
    monthly = _sum_metric(formulas.get("G"), "G", headers)
    metrics = {
        "daily": daily,
        "daily_rate": {"column": "F", "kind": "ratio", "source_metric": "daily", "denominator": "target", "formula": "daily_target_rate"},
        "original": monthly | {"column": "G"},
        "final": monthly | {"column": "H"},
        "sequential_rate": {"column": "I", "kind": "derived", "formula": "progressive_rate", "source_metric": "original", "denominator": "target", "total_days": 31, "elapsed_days_ref": "B3", "elapsed_days": 31},
        "rank": {"column": "J", "kind": "rank", "source_metric": "sequential_rate", "direction": "desc", "rank_ranges": rank_ranges},
        "product_daily": _sum_metric(formulas.get("K"), "K", headers, date_scope="day", product_filter=True),
        "product_monthly": _sum_metric(formulas.get("L"), "L", headers, product_filter=True),
    }
    return metrics


def _sum_metric(formula: str | None, column: str, headers: dict[str, str], date_scope: str | None = None, product_filter: bool = False) -> dict:
    text = str(formula or "")
    source_match = re.search(r"SUMIFS\([^!]+!\$(\w+):", text, re.IGNORECASE)
    source_column = source_match.group(1) if source_match else "W"
    metric = {"column": column, "source_column": source_column, "source_field": headers.get(source_column, source_column), "aggregate": "sum", "dimension_column": "BA", "dimension_field": headers.get("BA", "BA"), "filters": []}
    if date_scope:
        metric["date"] = {"field": headers.get("B", "B"), "column": "B", "scope": date_scope}
    if product_filter:
        metric["filters"] = [{"field": headers.get("C", "C"), "column": "C", "operator": "equals", "value": "升档专用合约"}]
    return metric
