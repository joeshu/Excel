from pathlib import Path
import re
from collections import Counter

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from app.services.formula_service import function_names


CELL_REFERENCE = re.compile(r"(?:(?:'([^']+)'|([A-Za-z0-9_][^!+\-*/(), ]*))!)?\$?([A-Z]{1,3})\$?(\d+)")


class TemplateParser:
    def parse(self, file_path: str) -> dict:
        workbook = load_workbook(file_path, read_only=False, data_only=False)
        sheets = []
        has_formula = False
        for worksheet in workbook.worksheets:
            columns = []
            for cell in worksheet[1]:
                if isinstance(cell, MergedCell):
                    continue
                values = []
                for row in worksheet.iter_rows(min_row=2, min_col=cell.column, max_col=cell.column):
                    value = row[0].value
                    if value is not None:
                        values.append(value)
                formula = None
                formula_row = None
                for row in range(2, worksheet.max_row + 1):
                    value = worksheet.cell(row=row, column=cell.column).value
                    if isinstance(value, str) and value.startswith("="):
                        formula = value
                        formula_row = row
                        break
                if formula:
                    has_formula = True
                columns.append({
                    "column": cell.column_letter,
                    "header": cell.value,
                    "type": self._infer_type(values, formula),
                    "is_formula": bool(formula),
                    "formula": formula,
                    "formula_row": formula_row,
                    "functions": function_names(formula) if formula else [],
                    "number_format": cell.number_format,
                    "hidden": bool(worksheet.column_dimensions[cell.column_letter].hidden),
                    "style_signature": self._style_signature(cell),
                    "formula_references": self._formula_references(formula, worksheet.title) if formula else [],
                })
            role_candidates = self._role_candidates(worksheet, workbook.sheetnames)
            is_detail_candidate = any(item["role"] == "detail" for item in role_candidates)
            field_header_row = self._field_header_row(worksheet) if is_detail_candidate else None
            sheets.append({
                "title": worksheet.title,
                "role_candidates": role_candidates,
                "max_row": worksheet.max_row,
                "max_column": worksheet.max_column,
                "header_rows": self._header_row_candidates(worksheet),
                "data_start_row_candidate": self._data_start_row_candidate(worksheet) if is_detail_candidate else None,
                "style_source_row_candidate": self._style_source_row_candidate(worksheet) if is_detail_candidate else None,
                "field_header_row_candidate": field_header_row,
                "field_headers": self._field_headers(worksheet, field_header_row) if field_header_row else [],
                "merged_ranges": [str(item) for item in worksheet.merged_cells.ranges],
                "hidden": worksheet.sheet_state != "visible",
                "freeze_panes": str(worksheet.freeze_panes) if worksheet.freeze_panes else None,
                "auto_filter": worksheet.auto_filter.ref,
                "print_area": str(worksheet.print_area) if worksheet.print_area else None,
                "print_title_rows": str(worksheet.print_title_rows) if worksheet.print_title_rows else None,
                "print_title_cols": str(worksheet.print_title_cols) if worksheet.print_title_cols else None,
                "hidden_rows": [row for row, dimension in worksheet.row_dimensions.items() if dimension.hidden],
                "formula_count": sum(1 for row in worksheet.iter_rows() for cell in row if self._is_formula(cell.value)),
                "formula_risks": self._formula_risks(worksheet),
                "columns": columns,
            })
        return {"sheets": sheets, "sheet_count": len(sheets), "has_formula": has_formula}

    @staticmethod
    def _infer_type(values: list, formula: str | None) -> str:
        if formula:
            return "formula"
        if any(isinstance(value, bool) for value in values):
            return "bool"
        if any(hasattr(value, "year") and hasattr(value, "month") for value in values):
            return "date"
        if values and all(isinstance(value, (int, float)) for value in values):
            return "number"
        return "text"

    @staticmethod
    def _style_signature(cell) -> str:
        return ":".join(str(value) for value in (
            cell.style_id,
            cell.number_format,
            cell.alignment.horizontal,
            cell.fill.fill_type,
            cell.fill.fgColor.rgb,
            cell.font.bold,
        ))

    @staticmethod
    def _formula_references(formula: str, default_sheet: str) -> list[dict]:
        references = []
        for quoted_sheet, plain_sheet, column, row_number in CELL_REFERENCE.findall(formula):
            references.append({"sheet": quoted_sheet or plain_sheet or default_sheet, "cell": f"{column.upper()}{row_number}"})
        return references

    @staticmethod
    def _is_formula(value) -> bool:
        return isinstance(value, str) and value.startswith("=")

    @staticmethod
    def _role_candidates(worksheet, sheet_names: list[str]) -> list[dict]:
        title = worksheet.title.lower()
        formula_count = sum(1 for row in worksheet.iter_rows() for cell in row if TemplateParser._is_formula(cell.value))
        text = " ".join(str(cell.value or "") for row in worksheet.iter_rows(min_row=1, max_row=min(worksheet.max_row, 10)) for cell in row)
        candidates = []
        if any(token in title for token in ("明细", "数据", "detail", "data")) or worksheet.max_row > 100:
            candidates.append({"role": "detail", "reason": "Sheet 名称或数据行数符合明细页特征"})
        if any(token in title for token in ("模版", "模板", "通报", "report")) or formula_count > 5:
            candidates.append({"role": "notice", "reason": "Sheet 名称或公式数量符合通报页特征"})
        if any(token in text for token in ("参数", "配置")) and worksheet.max_row < 50 and formula_count == 0:
            candidates.append({"role": "parameter", "reason": "内容包含参数或周期字段"})
        if not candidates:
            candidates.append({"role": "helper", "reason": "未匹配到明细页、通报页或参数页特征"})
        return candidates

    @staticmethod
    def _header_row_candidates(worksheet) -> list[dict]:
        candidates = []
        for row_number in range(1, min(worksheet.max_row, 20) + 1):
            values = [cell.value for cell in worksheet[row_number]]
            text_count = sum(1 for value in values if isinstance(value, str) and value.strip())
            nonempty_count = sum(value is not None for value in values)
            if text_count >= 2 and nonempty_count >= 2:
                candidates.append({"row": row_number, "text_count": text_count, "nonempty_count": nonempty_count})
        return candidates

    @staticmethod
    def _data_start_row_candidate(worksheet) -> int | None:
        for row_number in range(1, min(worksheet.max_row, 20) + 1):
            values = [cell.value for cell in worksheet[row_number]]
            if sum(value is not None for value in values) < 2:
                continue
            next_values = [cell.value for cell in worksheet[row_number + 1]] if row_number < worksheet.max_row else []
            if sum(value is not None for value in next_values) >= 2 and any(isinstance(value, (int, float)) for value in next_values):
                return row_number + 1
        return None

    @staticmethod
    def _style_source_row_candidate(worksheet) -> int | None:
        data_start = TemplateParser._data_start_row_candidate(worksheet)
        if data_start is None:
            return None
        return data_start

    @staticmethod
    def _field_header_row(worksheet) -> int | None:
        candidates = []
        for row_number in range(1, min(worksheet.max_row, 20) + 1):
            values = [cell.value for cell in worksheet[row_number]]
            text_count = sum(1 for value in values if isinstance(value, str) and value.strip())
            if text_count >= 2:
                candidates.append((text_count, row_number))
        return max(candidates, default=(0, None))[1]

    @staticmethod
    def _field_headers(worksheet, row_number: int | None) -> list[dict]:
        if row_number is None:
            return []
        return [{"column": cell.column_letter, "header": cell.value} for cell in worksheet[row_number] if cell.value is not None and str(cell.value).strip()]

    @staticmethod
    def _formula_risks(worksheet) -> list[dict]:
        risks = []
        formulas = [cell.value for row in worksheet.iter_rows() for cell in row if TemplateParser._is_formula(cell.value)]
        counts = Counter()
        for formula in formulas:
            if re.search(r"/[0-9]+", formula):
                counts["fixed_period_constant"] += 1
            if re.search(r"\$?[A-Z]{1,3}:\$?[A-Z]{1,3}", formula):
                counts["whole_column_reference"] += 1
            if re.search(r'"[^"\n]+"', formula):
                counts["hardcoded_dimension_value"] += 1
            if "!" in formula:
                counts["cross_sheet_reference"] += 1
            if re.search(r"RANK\([^,]+,\$?[A-Z]+\$?\d+:\$?[A-Z]+\$?\d+", formula, re.IGNORECASE):
                counts["fixed_rank_range"] += 1
        labels = {
            "fixed_period_constant": "公式包含固定周期常量",
            "whole_column_reference": "公式使用整列引用",
            "hardcoded_dimension_value": "公式包含硬编码维度值",
            "cross_sheet_reference": "公式包含跨 Sheet 引用",
            "fixed_rank_range": "排名公式使用固定范围",
        }
        for risk_type, count in counts.items():
            risks.append({"type": risk_type, "count": count, "message": labels[risk_type]})
        return risks


def ensure_xlsx(filename: str) -> None:
    if Path(filename).suffix.lower() != ".xlsx":
        raise ValueError("仅支持 .xlsx 文件")
