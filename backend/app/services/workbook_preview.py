from openpyxl import load_workbook


def preview_workbook(file_path: str, limit: int = 20) -> dict:
    workbook = load_workbook(file_path, read_only=True, data_only=False)
    sheets = []
    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(min_row=1, max_row=max(1, min(limit, worksheet.max_row)), values_only=True))
        sheets.append({
            "title": worksheet.title,
            "max_row": worksheet.max_row,
            "max_column": worksheet.max_column,
            "rows": rows,
        })
    return {"sheet_count": len(sheets), "sheets": sheets}
