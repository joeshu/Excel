from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook

from app.services.data_reader import read_records


def _value_type(value: object) -> str:
    if value is None or value == "":
        return "empty"
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, (datetime, date)):
        return "date"
    if isinstance(value, (int, float)):
        return "number"
    return "text"


def inspect_data_quality(file_path: str) -> dict:
    records = read_records(file_path)
    fields = list(records[0].keys()) if records else []
    issues = []
    field_stats = []
    for field in fields:
        values = [record.get(field) for record in records]
        non_empty = [value for value in values if _value_type(value) != "empty"]
        types = sorted({_value_type(value) for value in non_empty})
        missing_count = len(values) - len(non_empty)
        if missing_count:
            issues.append({"row": None, "field": field, "type": "missing_value", "message": f"字段 {field} 有 {missing_count} 个空值"})
        if len(types) > 1:
            issues.append({"row": None, "field": field, "type": "mixed_type", "message": f"字段 {field} 存在混合类型: {', '.join(types)}"})
        field_stats.append({"field": field, "rows": len(values), "missing": missing_count, "types": types})
    return {"row_count": len(records), "field_count": len(fields), "issue_count": len(issues), "valid": not issues, "fields": field_stats, "issues": issues}


def write_quality_report(file_path: str, output_path: str) -> str:
    report = inspect_data_quality(file_path)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "质量摘要"
    summary.append(["指标", "值"])
    summary.append(["数据行数", report["row_count"]])
    summary.append(["字段数", report["field_count"]])
    summary.append(["问题数", report["issue_count"]])
    summary.append(["状态", "通过" if report["valid"] else "存在问题"])
    fields = workbook.create_sheet("字段统计")
    fields.append(["字段", "行数", "空值数", "类型"])
    for item in report["fields"]:
        fields.append([item["field"], item["rows"], item["missing"], ", ".join(item["types"])])
    issues = workbook.create_sheet("问题明细")
    issues.append(["行号", "字段", "问题类型", "说明"])
    for item in report["issues"]:
        issues.append([item["row"] or "全列", item["field"], item["type"], item["message"]])
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path
