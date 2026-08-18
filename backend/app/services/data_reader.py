import csv
from pathlib import Path

from openpyxl import load_workbook


def read_records(file_path: str) -> list[dict]:
    suffix = Path(file_path).suffix.lower()
    if suffix == ".csv":
        with open(file_path, newline="", encoding="utf-8-sig") as source:
            return list(csv.DictReader(source))
    if suffix == ".xlsx":
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        headers = [str(value) if value is not None else "" for value in next(rows, ())]
        return [dict(zip(headers, row, strict=False)) for row in rows if any(value is not None for value in row)]
    raise ValueError("基础数据仅支持 .csv 或 .xlsx 文件")
