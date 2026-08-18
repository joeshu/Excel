from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.services.data_quality import inspect_data_quality


def append_final_sheets(output_path: str, records: list[dict], workflow, template, source) -> str:
    workbook = load_workbook(output_path)
    _remove_if_exists(workbook, "通报表")
    _remove_if_exists(workbook, "基础数据")
    _remove_if_exists(workbook, "工作流配置")
    _remove_if_exists(workbook, "数据质量报告")
    notice = workbook.create_sheet("通报表", 0)
    _write_notice(notice, records, template.name, workflow.name)
    source_sheet = workbook.create_sheet("基础数据")
    _write_records(source_sheet, records)
    config_sheet = workbook.create_sheet("工作流配置")
    _write_config(config_sheet, workflow, template, source)
    quality_sheet = workbook.create_sheet("数据质量报告")
    _write_quality(quality_sheet, source.file_path)
    workbook.save(output_path)
    return output_path


def _write_notice(sheet, records: list[dict], template_name: str, workflow_name: str) -> None:
    sheet.append(["Excel 通报表"])
    sheet.append(["模板", template_name])
    sheet.append(["工作流", workflow_name])
    sheet.append(["生成时间", datetime.utcnow()])
    sheet.append([])
    if not records:
        sheet.append(["结果", "没有可通报的数据"])
        _style_header(sheet)
        return
    headers = list(records[0].keys())
    sheet.append(headers)
    for record in records:
        sheet.append([record.get(header) for header in headers])
    _style_header(sheet, row=6)
    sheet.freeze_panes = "A7"
    sheet.auto_filter.ref = sheet.dimensions


def _write_records(sheet, records: list[dict]) -> None:
    headers = list(records[0].keys()) if records else []
    sheet.append(headers or ["结果"])
    for record in records:
        sheet.append([record.get(header) for header in headers])
    _style_header(sheet)
    sheet.freeze_panes = "A2"
    if headers:
        sheet.auto_filter.ref = sheet.dimensions


def _write_config(sheet, workflow, template, source) -> None:
    rows = [
        ["配置项", "值"],
        ["工作流名称", workflow.name],
        ["工作流模式", workflow.mode],
        ["模板名称", template.name],
        ["模板版本", template.version],
        ["数据源名称", source.name],
        ["数据源类型", source.source_type],
        ["数据源行数", _record_count(source.file_path)],
        ["字段映射", json.dumps(workflow.column_mapping or {}, ensure_ascii=False)],
        ["节点配置", json.dumps(workflow.node_json or {}, ensure_ascii=False)],
    ]
    for row in rows:
        sheet.append(row)
    _style_header(sheet)
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 100
    sheet.freeze_panes = "A2"


def _write_quality(sheet, file_path: str) -> None:
    report = inspect_data_quality(file_path)
    sheet.append(["指标", "值"])
    sheet.append(["数据行数", report["row_count"]])
    sheet.append(["字段数", report["field_count"]])
    sheet.append(["问题数", report["issue_count"]])
    sheet.append(["状态", "通过" if report["valid"] else "存在问题"])
    sheet.append([])
    sheet.append(["字段", "空值数", "类型"])
    for item in report["fields"]:
        sheet.append([item["field"], item["missing"], ", ".join(item["types"])])
    _style_header(sheet)
    _style_header(sheet, row=7)


def _record_count(file_path: str) -> int:
    from app.services.data_reader import read_records

    return len(read_records(file_path))


def _remove_if_exists(workbook, title: str) -> None:
    if title in workbook.sheetnames:
        del workbook[title]


def _style_header(sheet, row: int = 1) -> None:
    for cell in sheet[row]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor="17324D")
        cell.alignment = Alignment(horizontal="center")
