from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator


class WorkflowEngine:
    def __init__(self, template_path: str):
        self.workbook = load_workbook(template_path, read_only=False, data_only=False)

    def execute_formula_mode(self, data: list[dict], column_mapping: dict[str, str]):
        for sheet_index, worksheet in enumerate(self.workbook.worksheets):
            mappings = self._sheet_mappings(worksheet.title, column_mapping, sheet_index == 0)
            formula_rows = self._formula_rows(worksheet)
            for row_number, record in enumerate(data, start=2):
                for column, field in mappings.items():
                    cell = worksheet[f"{column}{row_number}"]
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        continue
                    self._write_preserving_style(cell, record.get(field))
                self._fill_formulas(worksheet, row_number, formula_rows)
        self.workbook.calculation.fullCalcOnLoad = True
        self.workbook.calculation.forceFullCalc = True
        return self.workbook

    @staticmethod
    def _sheet_mappings(sheet_title: str, mapping: dict[str, str], is_first_sheet: bool) -> dict[str, str]:
        qualified = {key.split("!", 1)[1]: value for key, value in mapping.items() if key.startswith(f"{sheet_title}!")}
        if qualified:
            return qualified
        if is_first_sheet:
            return {key: value for key, value in mapping.items() if "!" not in key and not key.startswith("__")}
        return {}

    @staticmethod
    def _formula_rows(worksheet) -> dict[int, int]:
        rows = {}
        for row in range(2, worksheet.max_row + 1):
            for column in range(1, worksheet.max_column + 1):
                value = worksheet.cell(row=row, column=column).value
                if isinstance(value, str) and value.startswith("="):
                    rows.setdefault(column, row)
        return rows

    def _fill_formulas(self, worksheet, row_number: int, formula_rows: dict[int, int]) -> None:
        for column in range(1, worksheet.max_column + 1):
            source_row = formula_rows.get(column)
            if source_row is None or row_number == source_row:
                continue
            source = worksheet.cell(row=source_row, column=column)
            target = worksheet.cell(row=row_number, column=column)
            if isinstance(source.value, str) and source.value.startswith("=") and target.value is None:
                target.value = self.translate_formula(source.value, source.coordinate, target.coordinate)
                target._style = copy(source._style)

    @staticmethod
    def _write_preserving_style(cell, value) -> None:
        style = copy(cell._style)
        cell.value = value
        cell._style = style

    def save(self, output_path: str) -> str:
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(output_path)
        return output_path

    @staticmethod
    def translate_formula(formula: str, origin: str, target: str) -> str:
        return Translator(formula, origin=origin).translate_formula(target)
